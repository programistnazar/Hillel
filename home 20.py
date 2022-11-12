# Прочитать сохранённый csv-файл из задания №18 и сохранить данные
# в excel-файл, кроме возраста – столбец с этими данными не нужен.

import csv
import openpyxl

with open('task_1.csv', 'r') as f:
    file_reader = list(csv.reader(f))

wb = openpyxl.Workbook()
wb.create_sheet(title='Первый лист', index=0)
wb.remove(wb['Sheet'])
sheet = wb['Первый лист']
for items, row in enumerate(file_reader):
    for col_index, value in enumerate(row):
        if col_index < 2:
            cell = sheet.cell(row=col_index+1, column=items+1)
            cell.value = value
        if col_index > 2:
            cell = sheet.cell(row=col_index, column=items+1)
            cell.value = value

wb.save('task_2.xlsx')


